import pandas as pd
from openpyxl import load_workbook

def analyze_all_sheets_structure(file_path):
    """5개 시트의 주차별 데이터 구조를 상세 분석"""
    
    wb = load_workbook(file_path, data_only=True)
    
    # 모든 시트 분석
    all_sheets = ['일편등심명동_0818', '육목원_0817', '대통령삼겹살 대학로점', '바다풍경2', '바다풍경1']
    
    for sheet_name in all_sheets:
        print(f"\n{'='*70}")
        print(f"시트명: {sheet_name}")
        print(f"{'='*70}")
        
        if sheet_name not in wb.sheetnames:
            print(f"❌ {sheet_name} 시트를 찾을 수 없습니다.")
            continue
            
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # 주차별 데이터 찾기
        weekly_data_found = False
        for i in range(len(df)):
            for j in range(len(df.columns)):
                cell_value = str(df.iloc[i, j])
                if '주차' in cell_value and ('1주차' in cell_value or '2주차' in cell_value):
                    weekly_data_found = True
                    print(f"\n주차별 데이터 발견: 행 {i}, '{cell_value}'")
                    
                    # 헤더 행도 출력
                    if i > 0:
                        header_row = df.iloc[i-1].tolist()
                        print(f"헤더 행 {i-1}: {header_row}")
                    
                    # 주차별 데이터 3개 행 출력
                    for k in range(i, min(i+6, len(df))):
                        row_data = df.iloc[k].tolist()
                        # NaN을 더 읽기 쉽게 표시
                        clean_row = []
                        for val in row_data:
                            if pd.isna(val):
                                clean_row.append("NaN")
                            else:
                                clean_row.append(str(val))
                        print(f"행 {k:2d}: {clean_row}")
                        
                        # 주차가 포함된 행인 경우 데이터 분석
                        if '주차' in str(df.iloc[k, j]):
                            print(f"    => 데이터 위치 분석:")
                            for idx, val in enumerate(row_data):
                                if not pd.isna(val) and val != '':
                                    try:
                                        num_val = float(val)
                                        if 0 < num_val < 1:  # CTR 같은 비율
                                            print(f"       위치 {idx}: {val} (CTR 후보)")
                                        elif 1000 <= num_val <= 50000:  # 노출수
                                            print(f"       위치 {idx}: {val} (노출수 후보)")
                                        elif 1 <= num_val <= 2000:  # 클릭수
                                            print(f"       위치 {idx}: {val} (클릭수 후보)")
                                        elif 1 <= num_val <= 10:  # 운영일수
                                            print(f"       위치 {idx}: {val} (운영일수 후보)")
                                    except:
                                        if isinstance(val, str) and ('주차' in val or '합계' in val):
                                            print(f"       위치 {idx}: {val} (레이블)")
                    print("-" * 50)
                    break
            if weekly_data_found:
                break
        
        if not weekly_data_found:
            print("❌ 주차별 데이터를 찾을 수 없습니다.")
            
        # 시트별 전체 구조 요약
        print(f"\n시트 전체 정보:")
        print(f"- 총 행 수: {len(df)}")
        print(f"- 총 열 수: {len(df.columns)}")

if __name__ == "__main__":
    file_path = "/Users/hongsukchoi/Desktop/cursor_project/01.live/cpc_monthly_report_웹통합_250817/0818_HTAG_CPC운영현황 보고서.xlsx"
    analyze_all_sheets_structure(file_path)
