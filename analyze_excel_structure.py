import pandas as pd
import openpyxl
from openpyxl import load_workbook

def analyze_excel_structure(file_path):
    """엑셀 파일의 구조를 분석하여 일자별 데이터 추출 범위를 확인"""
    
    print(f"=== {file_path} 분석 시작 ===")
    
    # openpyxl로 워크북 로드
    wb = load_workbook(file_path, data_only=True)
    
    # 모든 시트 이름 출력
    print(f"시트 목록: {wb.sheetnames}")
    
    # 호보식당_강남 시트 분석
    if "호보식당_강남" in wb.sheetnames:
        ws = wb["호보식당_강남"]
        print(f"\n=== 호보식당_강남 시트 분석 ===")
        
        # 시트의 전체 데이터 범위 확인
        print(f"시트 크기: {ws.max_row}행 x {ws.max_column}열")
        
        # 일자별 데이터 섹션 찾기
        daily_data_start_row = None
        daily_data_end_row = None
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and "[일자별 데이터]" in str(cell_value):
                    daily_data_start_row = row
                    print(f"일자별 데이터 헤더 발견: {row}행, {col}열")
                    break
            if daily_data_start_row:
                break
        
        if daily_data_start_row:
            print(f"\n=== 일자별 데이터 섹션 분석 (시작: {daily_data_start_row}행) ===")
            
            # 헤더 행 다음부터 데이터 확인
            data_start_row = daily_data_start_row + 2  # 헤더 + 컬럼명 행 건너뛰기
            
            print(f"데이터 시작 행: {data_start_row}")
            
            # 날짜 데이터 추출
            date_data = []
            for row in range(data_start_row, ws.max_row + 1):
                date_cell = ws.cell(row=row, column=2).value  # B열 (일자)
                if date_cell:
                    date_str = str(date_cell)
                    if "2025.08." in date_str:
                        spend = ws.cell(row=row, column=3).value  # C열 (소진비용)
                        exposure = ws.cell(row=row, column=4).value  # D열 (노출수)
                        clicks = ws.cell(row=row, column=5).value  # E열 (클릭수)
                        avg_cost = ws.cell(row=row, column=6).value  # F열 (클릭당평균비용)
                        
                        date_data.append({
                            'row': row,
                            'date': date_str,
                            'spend': spend,
                            'exposure': exposure,
                            'clicks': clicks,
                            'avg_cost': avg_cost
                        })
                        
                        print(f"{row}행: {date_str} | {spend} | {exposure} | {clicks} | {avg_cost}")
                else:
                    # 빈 셀을 만나면 데이터 끝으로 간주
                    if row > data_start_row + 5:  # 최소 5행은 확인
                        break
            
            print(f"\n=== 추출된 날짜 데이터 요약 ===")
            print(f"총 날짜 수: {len(date_data)}")
            if date_data:
                print(f"첫 번째 날짜: {date_data[0]['date']}")
                print(f"마지막 날짜: {date_data[-1]['date']}")
                
                # 8월 20일 이후 데이터 확인
                after_20th = [d for d in date_data if "2025.08.2" in d['date'] or "2025.08.3" in d['date']]
                print(f"8월 20일 이후 데이터: {len(after_20th)}개")
                for d in after_20th:
                    print(f"  {d['date']}: {d['spend']} | {d['exposure']} | {d['clicks']} | {d['avg_cost']}")
        else:
            print("일자별 데이터 헤더를 찾을 수 없습니다.")
    
    wb.close()

if __name__ == "__main__":
    analyze_excel_structure("HTAG_CPC운영현황 보고서_8월.xlsx")
